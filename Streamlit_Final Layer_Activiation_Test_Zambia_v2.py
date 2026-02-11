# -*- coding: utf-8 -*-
"""
Created on Wed Feb 11 10:27:23 2026

@author: tpriyank
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from streamlit_option_menu import option_menu


# =========================================================
# PAGE CONFIG
# =========================================================
favicon = "favicon.png"

st.set_page_config(
    page_title="LTE Layer Activation Processing Application",
    page_icon=favicon,
    layout="wide"
)

# =========================================================
# NOKIA STYLE SETTINGS
# =========================================================
background_text_color = "#001135"
background_header_text_color = "#a235b6"
background_font_style = "18px"
background_header_font_style = "22px"


# =========================================================
# GLOBAL CSS
# =========================================================
st.markdown("""
<style>
#MainMenu, footer, header {visibility: hidden;}

.stApp {
    background-color: white;
    font-family: "Nokia Pure Headline Light";
}

div.stButton > button {
    background-color: #a235b6;
    color: white;
    font-weight: bold;
    border-radius: 6px;
}
div.stButton > button:hover {
    background-color: #842b94;
}
</style>
""", unsafe_allow_html=True)


# =========================================================
# SIDEBAR (MATCHED TEMPLATE)
# =========================================================
with st.sidebar:

    selected = option_menu(
        menu_title="Airtel Zambia",
        options=["About", "Tool", "Contact Us"],
        icons=["person", "slack", "telephone"],
        styles={
            "container": {"background-color": "transparent"},
            "menu-title": {"color": "#660a93", "text-align": "center", "font-weight": "bold"},
            "nav-link": {
                "color": "#61206d",
                "font-size": "17px",
                "font-weight": "bold",
            },
            "nav-link-selected": {
                "background-color": "#a235b6",
                "color": "white"
            },
        },
    )


def get_band(cell):
    cell = str(cell).upper()
    if "L800" in cell: return "L800"
    if "L2600" in cell: return "L2600"
    if "L2100" in cell: return "L2100"
    return "L1800"


def get_sector(cell_name):
    m = re.search(r"S([1-9])", str(cell_name).upper())
    return f"S{m.group(1)}" if m else "unknown"


def get_carrier(cell_name):
    m = re.search(r"F([0-9]+)", str(cell_name).upper())
    return f"F{m.group(1)}" if m else "NA"


# =========================================================
# FULL PROCESS FUNCTION
# =========================================================
def run_full_analysis(bbh_file, day_file, sector_file):

    bbh_file.seek(0)
    day_file.seek(0)
    sector_file.seek(0)

    df_bbh = pd.read_excel(bbh_file,engine="openpyxl")
    df_day = pd.read_excel(day_file,engine="openpyxl")
    df_sector_input = pd.read_excel(sector_file,engine="openpyxl")

    # -------------------------------
    # Steps 1–3 (same as yours)
    # -------------------------------
    for df in (df_bbh, df_day, df_sector_input):
        df["Period start time"] = pd.to_datetime(df["Period start time"], errors="coerce")
        df["Date"] = df["Period start time"].dt.date
        df["Band"] = df["LNCEL name"].apply(get_band)
        df["Sector"] = df["LNCEL name"].apply(get_sector)
        df["Carrier"] = df["LNCEL name"].apply(get_carrier)
        df["BandCarrier"] = df["Band"] + "_" + df["Carrier"]
    
    df_sector_input["Band"] = df_sector_input["LNCEL name"].apply(get_band)
    df_sector_input["Sector"] = df_sector_input["LNCEL name"].apply(get_sector)
    df_sector_input["Carrier"] = df_sector_input["LNCEL name"].apply(get_carrier)
    df_sector_input["BandCarrier"] = df_sector_input.apply(lambda x: f"{x['Band']}_{x['Carrier']}" if x.get("Carrier","NA") not in (None,"NA","") else x["Band"], axis=1)
    
    
    if "Avg IP thp DL QCI9" in df_day.columns:
        df_day["Avg IP thp DL QCI9"] = pd.to_numeric(df_day["Avg IP thp DL QCI9"], errors="coerce")
    
    if "Avg IP thp DL QCI9" in df_bbh.columns:
        df_bbh["Avg IP thp DL QCI9"] = pd.to_numeric(df_bbh["Avg IP thp DL QCI9"], errors="coerce")
    
    
    if "Total LTE data volume, DL + UL" in df_day.columns:
        df_day.rename(columns={"Total LTE data volume, DL + UL": "Total LTE data volume, DL + UL (Daily)"}, inplace=True)
    
    # -------------------------------
    # Step 4: Sheet1 - BBH per LNBTS/Sector/Band (detailed)
    # -------------------------------
    print("Preparing Sheet1 (BBH detailed)...")
    bbh_kpis = [
        "Total E-UTRAN RRC conn stp SR",
        "E-UTRAN E-RAB stp SR","E-RAB DR RAN","Intra eNB HO SR",
        "inter eNB E-UTRAN HO SR X2","Avg RRC conn UE","Average CQI",
        "Avg UE distance","Total LTE data volume, DL + UL",
        "Avg IP thp DL QCI9","Avg PDCP cell thp DL",
        "RSSI_PUCCH_AVG (M8005C2)","Avg RSSI for PUSCH","SINR_PUCCH_AVG (M8005C92)",
        "SINR_PUSCH_AVG (M8005C95)","RACH Stp Completion SR",
        "Init Contx stp SR for CSFB","% MIMO RI 2","% MIMO RI 1",
        "Cell Avail excl BLU","E-UTRAN Avg PRB usage per TTI DL",
    ]
    
    bbh_kpis.append("Total LTE data volume, DL + UL (Daily)")
    
    records1 = []
    for kpi in bbh_kpis:
        if kpi in df_bbh.columns:
            temp = df_bbh.groupby(["LNBTS name","LNCEL name","Band","Sector","Date"])[kpi].mean().reset_index()
            temp = temp.melt(id_vars=["LNBTS name","LNCEL name","Band","Sector","Date"], value_vars=[kpi],
                             var_name="KPI", value_name="Value")
            records1.append(temp)
        elif kpi == "Total LTE data volume, DL + UL (Daily)" and "Total LTE data volume, DL + UL (Daily)" in df_day.columns:
            temp = df_day.groupby(["LNBTS name","LNCEL name","Band","Sector","Date"])["Total LTE data volume, DL + UL (Daily)"].sum().reset_index()
            temp["KPI"] = "Total LTE data volume, DL + UL (Daily)"
            temp = temp.rename(columns={"Total LTE data volume, DL + UL (Daily)": "Value"})
            records1.append(temp)
    
    if records1:
        sheet1 = pd.concat(records1, ignore_index=True)
        sheet1 = sheet1.pivot_table(index=["LNBTS name","LNCEL name","Band","Sector","KPI"], columns="Date", values="Value", aggfunc="first").reset_index()
    else:
        sheet1 = pd.DataFrame()
    
    # -------------------------------
    # Step 5: Sheet2 - Daily Average KPIs
    # -------------------------------
    print("Preparing Sheet2 (Daily averages)...")
    daily_bbh_kpis = [
        "Cell Avail excl BLU","E-UTRAN Avg PRB usage per TTI DL",
        "Total LTE data volume, DL + UL","Avg UE distance","Average CQI",
        "Avg RRC conn UE","Intra eNB HO SR","E-RAB DR RAN",
        "E-UTRAN E-RAB stp SR","Total E-UTRAN RRC conn stp SR",
        "Avg IP thp DL QCI9","Avg RRC conn UE"
    ]
    
    sum_kpis = [
        "Total LTE data volume, DL + UL",
        "Avg RRC conn UE"
    ]
    
    records2 = []
    
    
    for kpi in daily_bbh_kpis:
    
        if kpi in df_bbh.columns:
    
    
            agg_func = "sum" if kpi in sum_kpis else "mean"
    
            temp = (
                df_bbh
                .groupby("Date")[kpi]
                .agg(agg_func)
                .reset_index()
            )
    
            temp.rename(columns={kpi: "Value"}, inplace=True)
            temp["KPI"] = kpi
            records2.append(temp)
    
        else:
    
            temp = pd.DataFrame({
                "Date": df_bbh["Date"].unique(),
                "Value": np.nan,
                "KPI": kpi
            })
            records2.append(temp)
    
    if "Total LTE data volume, DL + UL (Daily)" in df_day.columns:
    
        payload_daily = (
            df_day
            .groupby("Date")["Total LTE data volume, DL + UL (Daily)"]
            .sum()
            .reset_index()
        )
    
        payload_daily.rename(
            columns={"Total LTE data volume, DL + UL (Daily)": "Value"},
            inplace=True
        )
    
        payload_daily["KPI"] = "Total LTE data volume, DL + UL (Daily)"
        records2.append(payload_daily)
    
    # ---------------------------------------------------
    # FINAL SHEET2
    # ---------------------------------------------------
    sheet2 = pd.concat(records2, ignore_index=True)
    
    sheet2 = (
        sheet2
        .pivot_table(
            index="KPI",
            columns="Date",
            values="Value",
            aggfunc="first"
        )
        .reset_index()
    )
    
    # -------------------------------
    # Step 6: Sheet3 - Band-level summary + % Traffic Distribution
    # -------------------------------
    print("Preparing Sheet3 (Band summary)...")
    
    kpis_sheet3 = [
        "Average CQI", "Avg RRC conn UE", "Avg UE distance",
        "Avg IP thp DL QCI9", "Cell Avail excl BLU", "E-RAB DR RAN",
        "E-UTRAN Avg PRB usage per TTI DL", "E-UTRAN E-RAB stp SR",
        "Intra eNB HO SR", "Total E-UTRAN RRC conn stp SR",
        "Total LTE data volume, DL + UL", "Total LTE data volume, DL + UL (Daily)"
    ]
    
    sum_kpis = {
        "Total LTE data volume, DL + UL",
        "Total LTE data volume, DL + UL (Daily)",
        "Avg RRC conn UE"
    }
    
    records3 = []
    
    
    df_combined = pd.concat([df_bbh, df_day], sort=False)
    
    if "Total LTE data volume, DL + UL (Daily)" in df_day.columns:
        for date, df_date in df_day.groupby("Date"):
            total_payload = df_date["Total LTE data volume, DL + UL (Daily)"].sum()
    
            for band, band_group in df_date.groupby("Band"):
                band_payload = band_group["Total LTE data volume, DL + UL (Daily)"].sum()
                pct_traffic = (band_payload / total_payload * 100) if total_payload > 0 else 0
    
                records3.append({
                    "KPI": "% Traffic Distribution",
                    "Band": band,
                    "Date": date,
                    "Value": pct_traffic
                })
    
    for kpi in kpis_sheet3:
    
        if kpi not in df_combined.columns:
            continue
    
        for (date, band), group in df_combined.groupby(["Date", "Band"]):
    
            if kpi in sum_kpis:
                value = group[kpi].sum()
            else:
                value = group[kpi].mean()
    
            records3.append({
                "KPI": kpi,
                "Band": band,
                "Date": date,
                "Value": value
            })
    
    sheet3 = pd.DataFrame(records3)
    
    if not sheet3.empty:
        sheet3 = (
            sheet3
            .pivot_table(
                index=["KPI", "Band"],
                columns="Date",
                values="Value",
                aggfunc="first"
            )
            .reset_index()
        )
    else:
        sheet3 = pd.DataFrame()
    
    print("Sheet3 (Band summary) prepared successfully")
    
    # -------------------------------
    # Step 7: Sheet4 - Config + Band Summary
    # -------------------------------
    print("Preparing Sheet4 (Config + Band summary)...")
    site_band_map = df_bbh.groupby("LNBTS name")["Band"].unique().apply(lambda x: "_".join(sorted(set(x)))).reset_index()
    site_band_map.rename(columns={"Band": "Configuration"}, inplace=True)
    
    df_bbh = df_bbh.merge(site_band_map, on="LNBTS name", how="left")
    df_day = df_day.merge(site_band_map, on="LNBTS name", how="left")
    df_combined_conf = pd.concat([df_bbh, df_day], sort=False)
    
    records4 = []
    
    if "Total LTE data volume, DL + UL (Daily)" in df_day.columns:
        for (date, conf), df_conf in df_day.groupby(["Date","Configuration"]):
            total_payload = df_conf["Total LTE data volume, DL + UL (Daily)"].sum()
            for band, band_group in df_conf.groupby("Band"):
                pct_traffic = (band_group["Total LTE data volume, DL + UL (Daily)"].sum() / total_payload * 100) if total_payload > 0 else 0
                records4.append({"KPI":"% Traffic Distribution","Configuration":conf,"Band":band,"Date":date,"Value":pct_traffic})
    
    
    for kpi in [k for k in kpis_sheet3 if k not in ["Total LTE data volume, DL + UL","Total LTE data volume, DL + UL (Daily)"]]:
        if kpi not in df_combined_conf.columns:
            continue
        for (date, conf, band), group in df_combined_conf.groupby(["Date","Configuration","Band"]):
            records4.append({"KPI":kpi,"Configuration":conf,"Band":band,"Date":date,"Value":group[kpi].mean()})
    
    
    for kpi in ["Total LTE data volume, DL + UL","Total LTE data volume, DL + UL (Daily)"]:
        for (date, conf, band), group in df_combined_conf.groupby(["Date","Configuration","Band"]):
            if kpi in group.columns:
                records4.append({"KPI":kpi,"Configuration":conf,"Band":band,"Date":date,"Value":group[kpi].sum()})
    
    sheet4 = pd.DataFrame(records4)
    if not sheet4.empty:
        sheet4 = sheet4.pivot_table(index=["KPI","Configuration","Band"], columns="Date", values="Value", aggfunc="first").reset_index()
    else:
        sheet4 = pd.DataFrame()
    
    # -------------------------------
    # Step 8: Sector-Band KPI summary
    # -------------------------------
    print("Preparing Sector-Band KPI summary (colored sheet)...")
    
    kpi_mapping = {
        "Avg IP thp DL QCI9": "Throughput",
        "E-UTRAN Avg PRB usage per TTI DL": "PRB",
        "Average CQI": "CQI",
        "Avg UE distance": "UEdist",
        "Total LTE data volume, DL + UL": "DataVol",
        "Avg RRC conn UE": "Users"
    }
    
    sum_kpis = [
        "Total LTE data volume, DL + UL",
        "Avg RRC conn UE"
    ]
    
    final_sector = (
        df_sector_input[["LNBTS name", "Sector"]]
        .drop_duplicates()
        .reset_index(drop=True)
    )
    
    for col, short in kpi_mapping.items():
    
        if col not in df_sector_input.columns:
            continue
    
        agg_func = "sum" if col in sum_kpis else "mean"
    
        summary = (
            df_sector_input
            .groupby(["LNBTS name", "Sector", "BandCarrier"])[col]
            .agg(agg_func)
            .reset_index()
        )
    
        pivot_summary = (
            summary
            .pivot_table(
                index=["LNBTS name", "Sector"],
                columns="BandCarrier",
                values=col,
                aggfunc="mean"  
            )
            .reset_index()
        )
    
        pivot_summary.columns = [
            f"{short}_{band}" if band not in ["LNBTS name", "Sector"] else band
            for band in pivot_summary.columns
        ]
    
        final_sector = pd.merge(
            final_sector,
            pivot_summary,
            on=["LNBTS name", "Sector"],
            how="left"
        )
    
    
    # -------------------------------
    # Step 9: Throughput analysis (Below_3/5, consecutive and 2-day avg )
    # -------------------------------
    print("Preparing throughput analysis sheets...")
    
    
    df_th = df_bbh.copy()
    if "Avg IP thp DL QCI9" in df_th.columns:
        df_th["DL_Mbps"] = pd.to_numeric(df_th["Avg IP thp DL QCI9"], errors="coerce")
    else:
        df_th["DL_Mbps"] = pd.NA
    
    def make_pivot_from_summary(summary_df, index_col="Band", col_col="Date", val_col="Value"):
        if summary_df.empty:
            return pd.DataFrame()
        pivot = summary_df.pivot(index=index_col, columns=col_col, values=val_col).fillna(0)
    
        try:
            pivot = pivot.astype(int)
        except Exception:
            pass
    
        if not pivot.empty:
            pivot = pivot.loc[:, (pivot != 0).any(axis=0)]
        return pivot
    
    
    def daily_counts(threshold):
        df_f = df_th[df_th["DL_Mbps"] < threshold]
        if df_f.empty:
            return pd.DataFrame()
        summary = df_f.groupby(["Band","Date"]).size().reset_index(name=f"Cells <{threshold}")
        return make_pivot_from_summary(summary, index_col="Band", col_col="Date", val_col=f"Cells <{threshold}")
    
    pivot_3mbps = daily_counts(3000)
    pivot_5mbps = daily_counts(5000)
    
    
    def consecutive_intersection(threshold):
        failing = df_th[df_th["DL_Mbps"] < threshold]
        results = []
        for band, group in failing.groupby("Band"):
            group = group.sort_values("Date")
            dates = sorted(group["Date"].unique())
            for i in range(1, len(dates)):
                d1, d2 = dates[i-1], dates[i]
                cells_d1 = set(group[group["Date"] == d1]["LNCEL name"])
                cells_d2 = set(group[group["Date"] == d2]["LNCEL name"])
                common = cells_d1 & cells_d2
                if common:
                    results.append({"Band": band, "Date": d2, "Value": len(common)})
        return make_pivot_from_summary(pd.DataFrame(results), index_col="Band", col_col="Date", val_col="Value")
    
    pivot_3mbps_2days = consecutive_intersection(3000)
    pivot_5mbps_2days = consecutive_intersection(5000)
    
    
    def avg_failures(threshold):
    
        cell_daily = df_th.groupby(["LNCEL name", "Band", "Date"])["DL_Mbps"].mean().reset_index()
        if cell_daily.empty:
            return pd.DataFrame()
        cell_daily = cell_daily.sort_values(["LNCEL name","Date"])
    
        cell_daily["2day_avg"] = cell_daily.groupby("LNCEL name")["DL_Mbps"].rolling(2).mean().reset_index(level=0, drop=True)
        failing = cell_daily[cell_daily["2day_avg"] < threshold].copy()
        if failing.empty:
            return pd.DataFrame()
        summary = failing.groupby(["Band","Date"]).size().reset_index(name=f"Cells <{threshold} (2day avg)")
        return make_pivot_from_summary(summary, index_col="Band", col_col="Date", val_col=f"Cells <{threshold} (2day avg)")
    
    pivot_3mbps_avg = avg_failures(3000)
    pivot_5mbps_avg = avg_failures(5000)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        sheet1.to_excel(writer, index=False, sheet_name="Sheet1_BBH_Detailed")
        sheet2.to_excel(writer, index=False, sheet_name="Sheet2_Daily_Avg")
        sheet3.to_excel(writer, index=False, sheet_name="Sheet3_Band_Summary")
        sheet4.to_excel(writer, index=False, sheet_name="Sheet4_Config_Band_Summary")
        final_sector.to_excel(writer, index=False, sheet_name="Sector_Band_KPIs")

        pivot_3mbps.to_excel(writer, sheet_name="Below_3Mbps")
        pivot_5mbps.to_excel(writer, sheet_name="Below_5Mbps")
        pivot_3mbps_2days.to_excel(writer, sheet_name="Below_3Mbps_2days")
        pivot_5mbps_2days.to_excel(writer, sheet_name="Below_5Mbps_2days")
        pivot_3mbps_avg.to_excel(writer, sheet_name="2day_Avg_3Mbps")
        pivot_5mbps_avg.to_excel(writer, sheet_name="2day_Avg_5Mbps")

    output.seek(0)

    # =====================================================
    # HEADER COLORING
    # =====================================================
    wb = load_workbook(output)

    if "Sector_Band_KPIs" in wb.sheetnames:
        ws = wb["Sector_Band_KPIs"]

        colors = {
            "Throughput": "FFFF00",
            "PRB": "00FF00",
            "CQI": "00B0F0",
            "UEdist": "FFC0CB",
            "DataVol": "FFA500",
            "Users": "C6EFCE"
        }

        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(1, col).value)
            for k, c in colors.items():
                if k in header:
                    fill = PatternFill(start_color=c, end_color=c, fill_type="solid")
                    for r in range(1, ws.max_row + 1):
                        ws.cell(r, col).fill = fill

    final = BytesIO()
    wb.save(final)
    final.seek(0)

    return final


if selected == "About":

    st.markdown(
        f"<h2 style='color:{background_header_text_color};'>LTE Layer Activation Tool</h2>",
        unsafe_allow_html=True
    )

    st.write("""
    This tool performs:
    • BBH KPI processing  
    • Daily aggregation  
    • Band summaries  
    • Configuration summaries  
    • Sector KPIs  
    • Throughput analysis (<3 / <5 Mbps)  
    • Automatic Excel generation  

    Upload 3 input files and download 11-sheet report automatically.
    """)

if selected == "Tool Name":

    st.markdown(
        f"<h2 style='color:{background_header_text_color};'>📊 LTE Layer Activation Tool</h2>",
        unsafe_allow_html=True
    )

    col1, col2, col3 = st.columns(3)

    with col1:
        bbh = st.file_uploader("Upload BBH Long", type="xlsx")

    with col2:
        day = st.file_uploader("Upload Daily File", type="xlsx")

    with col3:
        sector = st.file_uploader("Upload Sector File", type="xlsx")


    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("Run Full Analysis"):

        if not (bbh and day and sector):
            st.warning("⚠ Please upload all 3 files")
        else:
            with st.spinner("Processing 11 sheets..."):
                result = run_full_analysis(bbh, day, sector)

            st.success("✅ Processing Completed")

            st.download_button(
                "⬇ Download Output Excel",
                result,
                file_name="Output_Final_Combined.xlsx"
            )

if selected == "Contact Us":

    st.markdown(
        f"<h3 style='color:{background_header_text_color};'>Help us to improve!</h3>",
        unsafe_allow_html=True
    )

    st.write("For support or queries contact:")
    st.write("tomar.priyank@nokia.com")



