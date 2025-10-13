# -*- coding: utf-8 -*-
"""
Created on Tue Oct  7 15:54:20 2025

@author: tpriyank
"""

# streamlit_full_report.py
# -*- coding: utf-8 -*-
"""

"""
import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="LTE KPI Full Report", layout="wide")
st.title("📡 LTE KPI Full Report (Sheet1..SheetX + Throughput Analysis)")

st.markdown(
    "Upload the three Excel files below (BBH KPI, Day-level payload, Sector report). "
    "The app will produce a combined Excel workbook with all sheets and colored sector headers."
)

# File upload widgets
bbh_file = st.file_uploader("Upload BBH KPI file (e.g. Batch 2 Long Trend BBH KPI 7.xlsx)", type=["xlsx", "xls"])
day_file = st.file_uploader("Upload Day-level payload file (e.g. Batch 2 Day 6.xlsx)", type=["xlsx", "xls"])
sector_file = st.file_uploader("Upload Sector report file (optional, used for Sector_Band_KPIs)", type=["xlsx", "xls"])

# Helper functions
def get_band(cell_name):
    if pd.isna(cell_name): return "Unknown"
    try:
        match = re.search(r"\d+$", str(cell_name).strip())
        if match:
            num = int(match.group())
            if num in [11,12,13]: return "L2100"
            if num % 10 in [1,2,3]: return "L1800"
            if num % 10 in [4,5,6]: return "L2600"
            if num % 10 in [7,8,9]: return "L900"
        return "Other"
    except:
        return "Unknown"

def get_sector(cell_name):
    if pd.isna(cell_name): return "Sector-Unknown"
    try:
        match = re.search(r"\d+$", str(cell_name).strip())
        if match:
            suffix = int(match.group())
            if suffix in [1,11,4,7]: return "Sector-1"
            if suffix in [2,12,5,8]: return "Sector-2"
            if suffix in [3,13,6,9]: return "Sector-3"
        return "Sector-Other"
    except:
        return "Sector-Unknown"

def override_band_sector(lncel, band, sector):
    corrections = {
        "E_ANDOHARANOFOTSY-4": {"Band": "L1800", "Sector": "Sector-1"},
        "E_OTV-4": {"Band": "L1800", "Sector": "Sector-1"},
        "E_Ilakaka2-4": {"Band": "L1800", "Sector": "Sector-1"},
    }
    key = str(lncel)
    if key in corrections:
        return corrections[key]["Band"], corrections[key]["Sector"]
    return band, sector

def make_pivot(df_summary, index_col="Band", col_col="Date", val_col="Value"):
    if df_summary is None or df_summary.empty:
        return pd.DataFrame()
    pivot = df_summary.pivot(index=index_col, columns=col_col, values=val_col).fillna(0)
    # try convert to int if possible
    try:
        pivot = pivot.astype(int)
    except Exception:
        pass
    if not pivot.empty:
        pivot = pivot.loc[:, (pivot != 0).any(axis=0)]
    return pivot

# Run processing when all required files are uploaded
if bbh_file and day_file:
    st.success("BBH + Day files uploaded (Sector optional). Click **Run** to generate report.")
    if st.button("🚀 Run and generate full Excel report"):
        with st.spinner("Processing..."):
            # Read files into DataFrames
            df_bbh = pd.read_excel(bbh_file)
            df_day = pd.read_excel(day_file)
            # sector_file optional
            if sector_file:
                df_sector_input = pd.read_excel(sector_file)
            else:
                df_sector_input = pd.DataFrame()

            # Normalize Period start time -> Date
            for df in (df_bbh, df_day, df_sector_input):
                if "Period start time" in df.columns:
                    df["Period start time"] = pd.to_datetime(df["Period start time"], errors="coerce")
                    df["Date"] = df["Period start time"].dt.date

            # Apply band/sector and overrides to df_bbh and df_day and sector_input
            for df in (df_bbh, df_day):
                if "LNCEL name" not in df.columns:
                    st.error("One of the input files does not contain column 'LNCEL name'. Aborting.")
                    st.stop()
                df["Band"] = df["LNCEL name"].apply(get_band)
                df["Sector"] = df["LNCEL name"].apply(get_sector)
                df[["Band","Sector"]] = df.apply(
                    lambda x: override_band_sector(x["LNCEL name"], x["Band"], x["Sector"]),
                    axis=1, result_type="expand"
                )
                # ensure Date exists
                if "Date" not in df.columns:
                    df["Date"] = pd.NaT

            if not df_sector_input.empty and "LNCEL name" in df_sector_input.columns:
                df_sector_input["Band"] = df_sector_input["LNCEL name"].apply(get_band)
                df_sector_input["Sector"] = df_sector_input["LNCEL name"].apply(get_sector)
                df_sector_input[["Band","Sector"]] = df_sector_input.apply(
                    lambda x: override_band_sector(x["LNCEL name"], x["Band"], x["Sector"]),
                    axis=1, result_type="expand"
                )

            # Numeric conversions for throughput
            for df in (df_bbh, df_day):
                if "Non-GBR DL throughput" in df.columns:
                    df["Non-GBR DL throughput"] = pd.to_numeric(df["Non-GBR DL throughput"], errors="coerce")

            # ------------------------
            # Build Sheet1: BBH detailed
            # ------------------------
            st.info("Building Sheet1..Sheet4 and Sector_KPI...")
            bbh_kpis = [
                "Avg PRB usage per TTI UL","Avg IP thp DL QCI7","Avg IP thp DL QCI8",
                "Total E-UTRAN RRC conn stp SR2","Total E-UTRAN RRC conn stp SR",
                "E-UTRAN E-RAB stp SR","E-RAB DR RAN","Intra eNB HO SR",
                "inter eNB E-UTRAN HO SR X2","Avg RRC conn UE","Average CQI",
                "Avg UE distance","Total LTE data volume, DL + UL",
                "Avg IP thp DL QCI9","Avg PDCP cell thp DL","RRC_CONN_UE_MAX (M8001C200)",
                "RSSI_PUCCH_AVG (M8005C2)","Avg RSSI for PUSCH","SINR_PUCCH_AVG (M8005C92)",
                "SINR_PUSCH_AVG (M8005C95)","RACH Stp Completion SR",
                "Init Contx stp SR for CSFB","% MIMO RI 2","% MIMO RI 1",
                "Cell Avail excl BLU","E-UTRAN Avg PRB usage per TTI DL",
                "Non-GBR DL throughput","Non_GBR UL Throughput"
            ]
            bbh_kpis.append("Total LTE data volume, DL + UL (Daily)")

            records1 = []
            for kpi in bbh_kpis:
                if kpi in df_bbh.columns:
                    temp = df_bbh.groupby(["LNBTS name","LNCEL name","Band","Sector","Date"])[kpi].mean().reset_index()
                    temp = temp.melt(
                        id_vars=["LNBTS name","LNCEL name","Band","Sector","Date"],
                        value_vars=[kpi], var_name="KPI", value_name="Value"
                    )
                    records1.append(temp)
                elif kpi == "Total LTE data volume, DL + UL (Daily)" and "Total LTE data volume, DL + UL" in df_day.columns:
                    temp = df_day.groupby(["LNBTS name","LNCEL name","Band","Sector","Date"])["Total LTE data volume, DL + UL"].sum().reset_index()
                    temp["KPI"] = "Total LTE data volume, DL + UL (Daily)"
                    temp = temp.rename(columns={"Total LTE data volume, DL + UL": "Value"})
                    records1.append(temp)

            sheet1 = pd.concat(records1, ignore_index=True) if records1 else pd.DataFrame()
            if not sheet1.empty:
                sheet1 = sheet1.pivot_table(index=["LNBTS name","LNCEL name","Band","Sector","KPI"], columns="Date", values="Value", aggfunc="first").reset_index()

            # ------------------------
            # Sheet2: Daily avg KPIs
            # ------------------------
            daily_bbh_kpis = [
                "Cell Avail excl BLU","E-UTRAN Avg PRB usage per TTI DL",
                "Total LTE data volume, DL + UL","Avg UE distance","Average CQI",
                "Avg RRC conn UE","Intra eNB HO SR","E-RAB DR RAN",
                "E-UTRAN E-RAB stp SR","Total E-UTRAN RRC conn stp SR",
                "Non-GBR DL throughput","Non_GBR UL Throughput","Avg RRC conn UE"
            ]
            records2 = []
            for kpi in daily_bbh_kpis:
                if kpi in df_bbh.columns:
                    agg_func = "sum" if kpi == "Total LTE data volume, DL + UL" else "mean"
                    temp = df_bbh.groupby("Date")[kpi].agg(agg_func).reset_index().rename(columns={kpi:"Value"})
                    temp["KPI"] = kpi
                    records2.append(temp)
            if "Total LTE data volume, DL + UL" in df_day.columns:
                payload_daily = df_day.groupby("Date")["Total LTE data volume, DL + UL"].sum().reset_index().rename(columns={"Total LTE data volume, DL + UL":"Value"})
                payload_daily["KPI"] = "Total LTE data volume, DL + UL (Daily)"
                records2.append(payload_daily)
            sheet2 = pd.concat(records2, ignore_index=True) if records2 else pd.DataFrame()
            if not sheet2.empty:
                sheet2 = sheet2.pivot_table(index="KPI", columns="Date", values="Value", aggfunc="first").reset_index()

            # ------------------------
            # Sheet3: Band-level summary + % traffic
            # ------------------------
            kpis_sheet3 = [
                "Average CQI","Avg RRC conn UE","Avg UE distance",
                "Non-GBR DL throughput","Non_GBR UL Throughput","Cell Avail excl BLU","E-RAB DR RAN",
                "E-UTRAN Avg PRB usage per TTI DL","E-UTRAN E-RAB stp SR","Intra eNB HO SR",
                "Total E-UTRAN RRC conn stp SR","Total LTE data volume, DL + UL","Total LTE data volume, DL + UL (Daily)"
            ]
            records3 = []
            df_combined = pd.concat([df_bbh, df_day], sort=False)
            if "Total LTE data volume, DL + UL" in df_day.columns:
                # note: using daily payload column name existing in day file (original scripts sometimes renamed)
                for date, df_date in df_day.groupby("Date"):
                    total_payload = df_date["Total LTE data volume, DL + UL"].sum()
                    for band, band_group in df_date.groupby("Band"):
                        pct_traffic = (band_group["Total LTE data volume, DL + UL"].sum() / total_payload * 100) if total_payload > 0 else 0
                        records3.append({"KPI":"% Traffic Distribution","Band":band,"Date":date,"Value":pct_traffic})
            for kpi in [k for k in kpis_sheet3 if k not in ["Total LTE data volume, DL + UL","Total LTE data volume, DL + UL (Daily)"]]:
                if kpi not in df_combined.columns:
                    continue
                for (date, band), group in df_combined.groupby(["Date","Band"]):
                    records3.append({"KPI":kpi,"Band":band,"Date":date,"Value":group[kpi].mean()})
            for kpi in ["Total LTE data volume, DL + UL","Total LTE data volume, DL + UL (Daily)"]:
                for (date, band), group in df_combined.groupby(["Date","Band"]):
                    if kpi in group.columns:
                        records3.append({"KPI":kpi,"Band":band,"Date":date,"Value":group[kpi].sum()})
            sheet3 = pd.DataFrame(records3)
            if not sheet3.empty:
                sheet3 = sheet3.pivot_table(index=["KPI","Band"], columns="Date", values="Value", aggfunc="first").reset_index()

            # ------------------------
            # Sheet4: Config + Band summary
            # ------------------------
            site_band_map = df_bbh.groupby("LNBTS name")["Band"].unique().apply(lambda x: "_".join(sorted(set(x)))).reset_index().rename(columns={"Band":"Configuration"})
            df_bbh = df_bbh.merge(site_band_map, on="LNBTS name", how="left")
            df_day = df_day.merge(site_band_map, on="LNBTS name", how="left")
            df_combined_conf = pd.concat([df_bbh, df_day], sort=False)
            records4 = []
            if "Total LTE data volume, DL + UL" in df_day.columns:
                for (date, conf), df_conf in df_day.groupby(["Date","Configuration"]):
                    total_payload = df_conf["Total LTE data volume, DL + UL"].sum()
                    for band, band_group in df_conf.groupby("Band"):
                        pct_traffic = (band_group["Total LTE data volume, DL + UL"].sum() / total_payload * 100) if total_payload>0 else 0
                        records4.append({"KPI":"% Traffic Distribution","Configuration":conf,"Band":band,"Date":date,"Value":pct_traffic})
            for kpi in [k for k in kpis_sheet3 if k not in ["Total LTE data volume, DL + UL","Total LTE data volume, DL + UL (Daily)"]]:
                if kpi not in df_combined_conf.columns:
                    continue
                for (date, conf, band), group in df_combined_conf.groupby(["Date","Configuration","Band"]):
                    records4.append({"KPI":kpi,"Configuration":conf,"Band":band,"Date":date,"Value":group[kpi].mean()})
            for kpi in ["Total LTE data volume, DL + UL","Total LTE data volume, DL + UL (Daily)"]:
                for (date, conf, band), group in df_combined_conf.groupby(["Date","Configuration","Band"]):
                    if kpi in group.columns:
                        records4.append({"KPI":kpi,"Configuration":conf,"Band":band,"Date":date,"Value":group[kpi].sum()})
            sheet4 = pd.DataFrame(records4)
            if not sheet4.empty:
                sheet4 = sheet4.pivot_table(index=["KPI","Configuration","Band"], columns="Date", values="Value", aggfunc="first").reset_index()

            # ------------------------
            # Sector-Band KPI summary (if sector file provided)
            # ------------------------
            final_sector = pd.DataFrame()
            if not df_sector_input.empty:
                kpi_mapping = {
                    "Non-GBR DL throughput": "Throughput",
                    "E-UTRAN Avg PRB usage per TTI DL": "PRB",
                    "Average CQI": "CQI",
                    "Avg UE distance": "UEdist",
                    "Total LTE data volume, DL + UL": "DataVol",
                    "Avg RRC conn UE": "Users"
                }
                final_sector = df_sector_input[["LNBTS name","Sector"]].drop_duplicates().reset_index(drop=True)
                for col, short in kpi_mapping.items():
                    if col not in df_sector_input.columns:
                        continue
                    agg_func = "sum" if col == "Total LTE data volume, DL + UL" else "mean"
                    summary = df_sector_input.groupby(["LNBTS name","Sector","Band"])[col].agg(agg_func).reset_index()
                    pivot_summary = summary.pivot_table(index=["LNBTS name","Sector"], columns="Band", values=col, aggfunc="mean").reset_index()
                    pivot_summary.columns = [f"{short}_{band}" if band not in ["LNBTS name","Sector"] else band for band in pivot_summary.columns]
                    final_sector = pd.merge(final_sector, pivot_summary, on=["LNBTS name","Sector"], how="left")

            # ------------------------
            # Throughput analysis (Band & Sector)
            # ------------------------
            df_th = df_day.copy() if not df_day.empty else df_bbh.copy()
            if "Non-GBR DL throughput" in df_th.columns:
                df_th["DL_Mbps"] = pd.to_numeric(df_th["Non-GBR DL throughput"], errors="coerce")
            else:
                df_th["DL_Mbps"] = pd.NA

            # Daily counts by grouping (Band / Sector)
            def daily_counts_group(threshold, group_by="Band"):
                df_f = df_th[df_th["DL_Mbps"] < threshold].copy()
                if df_f.empty: return pd.DataFrame()
                summary = df_f.groupby([group_by, "Date"])["LNCEL name"].nunique().reset_index(name=f"Cells <{threshold}")
                return make_pivot(summary, index_col=group_by, col_col="Date", val_col=f"Cells <{threshold}")

            # Consecutive intersection
            def consecutive_intersection_group(threshold, group_by="Band"):
                failing = df_th[df_th["DL_Mbps"] < threshold].copy()
                results = []
                if failing.empty: return pd.DataFrame()
                for grp_val, group in failing.groupby(group_by):
                    group = group.sort_values("Date")
                    dates = sorted(group["Date"].unique())
                    for i in range(1, len(dates)):
                        d1, d2 = dates[i-1], dates[i]
                        set_d1 = set(group[group["Date"]==d1]["LNCEL name"])
                        set_d2 = set(group[group["Date"]==d2]["LNCEL name"])
                        common = set_d1 & set_d2
                        if common:
                            results.append({group_by: grp_val, "Date": d2, "Value": len(common)})
                if not results: return pd.DataFrame()
                df_res = pd.DataFrame(results)
                return make_pivot(df_res, index_col=group_by, col_col="Date", val_col="Value")

            # 2-day avg per cell then count per group
            def avg_failures_group(threshold, group_by="Band"):
                cell_daily = df_th.groupby(["LNCEL name", group_by, "Date"])["DL_Mbps"].mean().reset_index()
                if cell_daily.empty: return pd.DataFrame()
                cell_daily = cell_daily.sort_values(["LNCEL name","Date"])
                cell_daily["2day_avg"] = cell_daily.groupby("LNCEL name")["DL_Mbps"].rolling(2).mean().reset_index(level=0, drop=True)
                failing = cell_daily[cell_daily["2day_avg"] < threshold].copy()
                if failing.empty: return pd.DataFrame()
                summary = failing.groupby([group_by, "Date"])["LNCEL name"].nunique().reset_index(name=f"Cells <{threshold} (2day avg)")
                return make_pivot(summary, index_col=group_by, col_col="Date", val_col=f"Cells <{threshold} (2day avg)")

            # compute for Band & Sector, thresholds 3000 & 5000
            pivot_3mbps_band = daily_counts_group(3000, "Band")
            pivot_5mbps_band = daily_counts_group(5000, "Band")
            pivot_3mbps_sector = daily_counts_group(3000, "Sector")
            pivot_5mbps_sector = daily_counts_group(5000, "Sector")

            pivot_3mbps_2days_band = consecutive_intersection_group(3000, "Band")
            pivot_5mbps_2days_band = consecutive_intersection_group(5000, "Band")
            pivot_3mbps_2days_sector = consecutive_intersection_group(3000, "Sector")
            pivot_5mbps_2days_sector = consecutive_intersection_group(5000, "Sector")

            pivot_3mbps_avg_band = avg_failures_group(3000, "Band")
            pivot_5mbps_avg_band = avg_failures_group(5000, "Band")
            pivot_3mbps_avg_sector = avg_failures_group(3000, "Sector")
            pivot_5mbps_avg_sector = avg_failures_group(5000, "Sector")

            # ------------------------
            # Write results to Excel in-memory, then color header for sector sheet
            # ------------------------
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Sheet1..4
                (sheet1.to_excel(writer, sheet_name="Sheet1_BBH_Detailed", index=False) if not sheet1.empty else pd.DataFrame().to_excel(writer, sheet_name="Sheet1_BBH_Detailed", index=False))
                (sheet2.to_excel(writer, sheet_name="Sheet2_Daily_Avg", index=False) if not sheet2.empty else pd.DataFrame().to_excel(writer, sheet_name="Sheet2_Daily_Avg", index=False))
                (sheet3.to_excel(writer, sheet_name="Sheet3_Band_Summary", index=False) if (not sheet3.empty) else pd.DataFrame().to_excel(writer, sheet_name="Sheet3_Band_Summary", index=False))
                (sheet4.to_excel(writer, sheet_name="Sheet4_Config_Band_Summary", index=False) if (not sheet4.empty) else pd.DataFrame().to_excel(writer, sheet_name="Sheet4_Config_Band_Summary", index=False))

                # Sector_Band_KPIs
                if not final_sector.empty:
                    final_sector.to_excel(writer, sheet_name="Sector_Band_KPIs", index=False)
                else:
                    pd.DataFrame().to_excel(writer, sheet_name="Sector_Band_KPIs", index=False)

                # Band-level throughput sheets
                (pivot_3mbps_band.to_excel(writer, sheet_name="Below_3Mbps_Band") if not pivot_3mbps_band.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_3Mbps_Band", index=False))
                (pivot_5mbps_band.to_excel(writer, sheet_name="Below_5Mbps_Band") if not pivot_5mbps_band.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_5Mbps_Band", index=False))
                (pivot_3mbps_2days_band.to_excel(writer, sheet_name="Below_3Mbps_2days_Band") if not pivot_3mbps_2days_band.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_3Mbps_2days_Band", index=False))
                (pivot_5mbps_2days_band.to_excel(writer, sheet_name="Below_5Mbps_2days_Band") if not pivot_5mbps_2days_band.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_5Mbps_2days_Band", index=False))
                (pivot_3mbps_avg_band.to_excel(writer, sheet_name="2dayAvg_3Mbps_Band") if not pivot_3mbps_avg_band.empty else pd.DataFrame().to_excel(writer, sheet_name="2dayAvg_3Mbps_Band", index=False))
                (pivot_5mbps_avg_band.to_excel(writer, sheet_name="2dayAvg_5Mbps_Band") if not pivot_5mbps_avg_band.empty else pd.DataFrame().to_excel(writer, sheet_name="2dayAvg_5Mbps_Band", index=False))

                # Sector-level throughput sheets
                (pivot_3mbps_sector.to_excel(writer, sheet_name="Below_3Mbps_Sector") if not pivot_3mbps_sector.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_3Mbps_Sector", index=False))
                (pivot_5mbps_sector.to_excel(writer, sheet_name="Below_5Mbps_Sector") if not pivot_5mbps_sector.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_5Mbps_Sector", index=False))
                (pivot_3mbps_2days_sector.to_excel(writer, sheet_name="Below_3Mbps_2days_Sector") if not pivot_3mbps_2days_sector.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_3Mbps_2days_Sector", index=False))
                (pivot_5mbps_2days_sector.to_excel(writer, sheet_name="Below_5Mbps_2days_Sector") if not pivot_5mbps_2days_sector.empty else pd.DataFrame().to_excel(writer, sheet_name="Below_5Mbps_2days_Sector", index=False))
                (pivot_3mbps_avg_sector.to_excel(writer, sheet_name="2dayAvg_3Mbps_Sector") if not pivot_3mbps_avg_sector.empty else pd.DataFrame().to_excel(writer, sheet_name="2dayAvg_3Mbps_Sector", index=False))
                (pivot_5mbps_avg_sector.to_excel(writer, sheet_name="2dayAvg_5Mbps_Sector") if not pivot_5mbps_avg_sector.empty else pd.DataFrame().to_excel(writer, sheet_name="2dayAvg_5Mbps_Sector", index=False))

            # We must apply colors to the Sector_Band_KPIs sheet headers inside the Excel file.
            # To do that, load the workbook from the BytesIO buffer, color, then write back to a new BytesIO.
            output.seek(0)
            wb = load_workbook(output)
            if "Sector_Band_KPIs" in wb.sheetnames:
                ws = wb["Sector_Band_KPIs"]
                kpi_colors = {
                    "Throughput": "FFFF00",  # yellow
                    "PRB": "00FF00",         # green
                    "CQI": "00B0F0",         # blue
                    "UEdist": "FFC0CB",      # pink
                    "DataVol": "FFA500",     # orange
                    "Users": "C6EFCE"        # light green
                }
                for col_idx in range(1, ws.max_column + 1):
                    header_value = ws.cell(row=1, column=col_idx).value
                    if not header_value:
                        continue
                    for kpi_key, color in kpi_colors.items():
                        if kpi_key in str(header_value):
                            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            for r in range(1, ws.max_row + 1):
                                ws.cell(row=r, column=col_idx).fill = fill
                            break
                # Save colored workbook into new BytesIO
                out2 = BytesIO()
                wb.save(out2)
                out2.seek(0)
            else:
                # No sector sheet or no coloring required
                out2 = BytesIO()
                wb.save(out2)
                out2.seek(0)

            st.success("Report generation complete.")
            st.download_button(
                label="📥 Download combined Excel report",
                data=out2.getvalue(),
                file_name="Final_Combined_Batch2_FullReport.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # quick preview tables for sanity checking
            st.subheader("Preview: Band-level daily counts (<3Mbps)")
            if not pivot_3mbps_band.empty:
                st.dataframe(pivot_3mbps_band.head(50))
            else:
                st.write("No Band-level <3Mbps counts found.")

            st.subheader("Preview: Sector_Band_KPIs (first 10 rows)")
            if not final_sector.empty:
                st.dataframe(final_sector.head(10))
            else:
                st.write("No Sector_Band_KPIs available (sector file missing or empty).")

else:
    st.info("Please upload BBH and Day files to enable processing. Sector file is optional.")
